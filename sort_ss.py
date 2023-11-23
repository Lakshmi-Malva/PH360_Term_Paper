from openpyxl import load_workbook, Workbook
import xlsxwriter as xlsxwt
import string
import pandas as pd
import numpy as np
import os
from Utility_funcs import *

'''Extracting node patterns of stable states'''

#Get the number of stable states saved
def get_ss_num(ss_file, network):
    path = os.getcwd() + "/OUTPUT/" + network +'/'
    sl_no = pd.read_excel(path+ss_file, sheet_name=network,usecols='A')
    sl_no = sl_no.to_numpy(); temp = []
    for i in range(len(sl_no)): temp.append(int(sl_no[i][0]))
    sl_no = np.array(temp)
    return int(len(sl_no))

#Create an array of the node patterns
def get_color_arr(ss_file,network,node_num):
    path = os.getcwd() + "/OUTPUT/" + network +'/'
    try: wb = load_workbook(path+ss_file,data_only = True)
    except: 
        wb = Workbook()
        wb.save(path+ss_file)
    try: sh = wb[network]
    except: 
        sh = wb.create_sheet(network)
        wb.save(path+ss_file)

    alphabets = [letter for letter in string.ascii_uppercase]
    row_label = alphabets[1:node_num+1]
    if node_num>25:  
        if node_num>=51: temp = alphabets
        else: temp = alphabets[:node_num-25]
        temp = ['A'+a for a in temp]
        row_label = row_label+temp
    if node_num>51:  
        temp = alphabets[:node_num-51]
        temp = ['B'+a for a in temp]
        row_label = row_label+temp

    ss_num = get_ss_num(ss_file, network)
    
    Color = []
    for j in range(ss_num):
        Color_hex = []
        for i in range(node_num):
            cell = row_label[i]+str(j+2)
            color_in_hex = sh[cell].fill.bgColor.index
            Color_hex.append(color_in_hex)
        Color.append(Color_hex)

    color_arr = []
    for i in range(len(Color)):
        lst = Color[i]
        c_lst = []
        for j in range(len(lst)):
            if lst[j] == 'FF000000': c_lst.append(1) #black
            elif lst[j] == '00000000': c_lst.append(-1) #white
        color_arr.append(c_lst)
    
    return np.array(color_arr), row_label

def node_sort(input,network,file,nodes,SteadyState,frustration):

    ss_file = input['ss_file']
    workbook = xlsxwt.Workbook(os.path.join(os.getcwd() + "/OUTPUT/" + network,file))
    worksheet = workbook.add_worksheet("stable_states")
    cell_format = workbook.add_format()
    cell_format.set_bg_color('black')

    for i, node in enumerate(nodes):
        worksheet.write(0, i+1, node) #Writing names of the nodes

    node_num = len(nodes)
    worksheet.write(0, node_num+2,'Frequency')
    worksheet.write(0, node_num+3,'Frustration')

    color_arr, row_label = get_color_arr(ss_file,network,node_num)
    init_clr_len = len(color_arr)
    
    def find_vect(vect,SteadyState):
        state_list = [list(str2vect(key,node_num)) for key in list(SteadyState.keys())]
        try: index_1 = state_list.index(list(vect))
        except: index_1 = None
        try: index_2 = state_list.index(list(-vect))
        except: index_2 = None
        if vect[0] == 1:
            vect_type = 'Black'
            index_b = index_1; index_w = index_2
        if vect[0] == -1: 
            vect_type = 'White'
            index_w = index_1; index_b = index_2
        return index_w, index_b, vect_type
    
    def update_out(input,index,index_c,vect,node_num,sort_SteadyState,frustration,worksheet,cell_format):
        for j,node in enumerate(vect): 
            if node == input['node_values'][0]:
                worksheet.write(index+1,j+1,str(node),cell_format)
                worksheet.write(index_c+1,j+1,str(-node))
            else: 
                worksheet.write(index+1,j+1,str(node))
                worksheet.write(index_c+1,j+1,str(-node),cell_format)
        
        state_c = vect2str(-vect,node_num)
        worksheet.write(index+1,node_num+2,str(values))
        worksheet.write(index_c+1,node_num+2,str(sort_SteadyState.get(state_c,0)))
        worksheet.write(index+1,node_num+3,str(frustration[state]))
        worksheet.write(index_c+1,node_num+3,str(frustration.get(state_c,0)))
        return

    sort_SteadyState = dict(sorted(SteadyState.items(), key=lambda x:x[1],reverse=True))
    Index_done = []
    color_arr = [col.tolist() for col in color_arr]

    for i,(state,values) in enumerate(sort_SteadyState.items()):
        vect = str2vect(state,node_num)
        index_w, index_b, vect_type = find_vect(vect,sort_SteadyState)
        proceed = any(i == ind for ind in Index_done)
        if proceed: continue
        
        for col in color_arr:
            compare = np.all([col[j] == vect[j] for j in range(len(col))])
            if compare: 
                ind = color_arr.index(col)
                break
        if init_clr_len == 0: compare = False

        if int(compare) == 0:
            if vect_type == 'Black': 
                color_arr.append(-vect); color_arr.append(vect)
                ind = len(color_arr) + 1
            if vect_type == 'White':
                color_arr.append(vect); color_arr.append(-vect)
                ind = len(color_arr) 
        
        if vect_type == 'Black' and index_b != None:
            index = ind; index_c = ind - 1 
        if vect_type == 'White' and index_w != None: 
            index = ind; index_c = ind + 1
        
        update_out(input,index,index_c,vect,node_num,sort_SteadyState,frustration,worksheet,cell_format)

        if index_w!=None: Index_done.append(index_w)
        if index_b!=None: Index_done.append(index_b)

    workbook.close()
    return color_arr,init_clr_len,row_label

def update(input,color_arr,init_clr_len,network,row_label):

    ss_file = input['ss_file']
    path = os.getcwd() + "/OUTPUT/" + network +'/'
    if len(color_arr) > init_clr_len:
        from openpyxl.styles import PatternFill 

        wb = load_workbook(path+ss_file,data_only = True)
        sh = wb[network]

        diff = len(color_arr)-init_clr_len
        for i in range(diff):
            row_num = init_clr_len+i+2
            sh['A'+str(row_num)].value = row_num-1
            for j,node in enumerate(color_arr[init_clr_len+i]):
                label = row_label[j]+str(row_num)
                if node == input['node_values'][0]:
                    sh[label].fill = PatternFill(bgColor="FF000000", fill_type = "solid")
                else: pass
        wb.save(path+ss_file)
        print('Stable states updated')
    return
